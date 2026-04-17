# app/service/excel/asset_import_service.py
import io
from typing import List, Dict, Any
from datetime import date, datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from app.schemas.assets.AssetExcel import AssetImportRow

def _get_excel_columns() -> List[tuple]:
    """Возвращает список колонок (field_name, header_name, category_color)"""
    # Цвета категорий для визуального разделения в Excel
    COLORS = {
        "main": "4472C4",      # Синий
        "location": "70AD47",  # Зеленый
        "details": "C55A11",   # Оранжевый
        "vendors": "9F4826",   # Коричневый
        "users": "7B7B7B",     # Серый
        "other": "A5A5A5"      # Светло-серый
    }

    schema_fields = AssetImportRow.model_fields
    columns = []

    # Маппинг полей на категории (упрощенно по префиксу или явный список)
    categories = {
        "inventory_id": "main", "serial_number": "main", "name": "main", "asset_type_name": "main", "asset_status": "main",
        "location_country": "location", "location_city": "location", "location_address": "location", "location_room": "location", "location_floor": "location",
        "type_domain": "details", "passwork": "details", "date_issue": "details", "date_purchasing": "details", "price": "details", "comment": "details", "affixed_inventory_id": "details",
        "manufacturer_name": "vendors", "vendor_name": "vendors",
        "prepared_by_name": "users", "checked_by_name": "users",
        "software_os_type": "other", "parent_inventory_id": "other"
    }

    for name, field in schema_fields.items():
        header = field.description if field.description else name.replace("_", " ").title()
        cat = categories.get(name, "other")
        columns.append((name, header, COLORS.get(cat, "A5A5A5")))

    return columns

def create_asset_import_template() -> bytes:
    """Генерирует шаблон Excel файла с цветными заголовками"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    columns = _get_excel_columns()

    for col_idx, (_, header, color) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # Пример данных
    example = [
        "INV-001", "SN-123", "Ноутбук Lenovo", "Ноутбук", "Приемка",
        "Россия", "Москва", "Ленина 1", "101", "1",
        "CORP", "admin", "2026-01-01", "2025-12-01", 50000, "Comment", "Да",
        "Lenovo", "DNS",
        "Иванов И.И.", "Петров П.П.",
        "Windows 11", ""
    ]
    for i, val in enumerate(example, 1):
        ws.cell(row=2, column=i, value=val)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


async def parse_asset_import_excel(file_content: bytes) -> List[Dict[str, Any]]:
    """Парсит Excel файл в список словарей"""
    df = pd.read_excel(io.BytesIO(file_content))

    rows = []
    # Создаем маппинг Заголовок -> ИмяПоля
    header_to_field = {header: field for field, header, _ in _get_excel_columns()}

    for _, row in df.iterrows():
        raw_dict = row.to_dict()
        mapped_dict = {}

        for excel_header, value in raw_dict.items():
            if excel_header in header_to_field:
                field_name = header_to_field[excel_header]

                # Обработка дат
                if field_name.endswith('_date'):
                    if pd.notna(value):
                        if isinstance(value, pd.Timestamp):
                            value = value.date()
                        elif isinstance(value, date):
                            pass
                        else:
                            try:
                                value = datetime.strptime(str(value), "%Y-%m-%d").date()
                            except:
                                value = None
                    else:
                        value = None

                # Обработка булевых значений
                if field_name == 'affixed_inventory_id':
                    if isinstance(value, bool):
                        pass
                    elif isinstance(value, str):
                        value = value.lower() in ['да', 'yes', 'true', '1']
                    else:
                        value = bool(value) if pd.notna(value) else False

                mapped_dict[field_name] = value

        # Валидация обязательных полей
        if not mapped_dict.get('inventory_id') or not mapped_dict.get('serial_number') or not mapped_dict.get('name') or not mapped_dict.get('asset_type_name'):
            continue

        rows.append(mapped_dict)

    return rows