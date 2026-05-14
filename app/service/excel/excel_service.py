import io
import math
from typing import List, Dict, Any
from datetime import date, datetime
import pandas as pd
from numpy import bool, dtype, ndarray
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pandas.core.generic import NDFrame

from app.schemas.assets.AssetExcel import AssetExcelRow

def _get_excel_columns() -> List[tuple]:
    """Возвращает список колонок (field_name, header_name, category_color)"""
    # Цвета категорий для визуального разделения в Excel
    COLORS = {
        "main": "4472C4",      # Синий
        "location": "70AD47",  # Зеленый
        "details": "C55A11",   # Оранжевый
        "vendors": "9F4826",   # Коричневый
        "users": "7B7B7B",     # Серый
        "other": "A5A5A5"      # Светло-серый
    }

    schema_fields = AssetExcelRow.model_fields
    columns = []

    # Маппинг полей на категории (упрощенно по префиксу или явный список)
    categories = {
        "inventory_id": "main", "serial_number": "main", "name": "main", "asset_type_name": "main", "asset_status": "main",
        "location_country": "location", "location_city": "location", "location_address": "location", "location_room": "location", "location_floor": "location",
        "type_domain": "details", "date_issue": "details", "date_purchasing": "details", "price": "details", "comment": "details", "affixed_inventory_id": "details",
        "manufacturer_name": "vendors", "vendor_name": "vendors",
        "prepared_by_name": "users", "checked_by_name": "users",
        "software_os_type": "other", "parent_inventory_id": "other"
    }

    for name, field in schema_fields.items():
        header = field.description if field.description else name.replace("_", " ").title()
        cat = categories.get(name, "other")
        columns.append((name, header, COLORS.get(cat, "A5A5A5")))

    return columns

def create_asset_export_excel(assets_data: List[Dict[str, Any]]) -> bytes:
    """
    Создает Excel файл для экспорта активов.
    assets_ Список словарей, где ключи соответствуют полям AssetExcelRow.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets Export"

    # Стили
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    columns = _get_excel_columns()

    # Заголовки
    for col_idx, (_, header, color) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.border = thin_border

    # Данные
    for row_idx, item in enumerate(assets_data, start=2):
        for col_idx, (field_name, _, _) in enumerate(columns, start=1):
            value = item.get(field_name, "")

            # Форматирование
            if isinstance(value, (date, datetime)):
                value = value.strftime("%Y-%m-%d")
            elif isinstance(value, bool):
                value = "Да" if value else "Нет"
            elif value is None:
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Автоширина
    for col_idx, (_, header, _) in enumerate(columns, start=1):
        max_length = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def create_asset_import_template() -> bytes:
    """Генерирует шаблон Excel файла с цветными заголовками"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    columns = _get_excel_columns()

    for col_idx, (_, header, color) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # Пример данных
    example = [
        "INV-001", "SN-123", "Ноутбук Lenovo", "Ноутбук", "Приемка",
        "Россия", "Москва", "Ленина 1", "101", "1",
        "CORP", "admin", "2026-01-01", "2025-12-01", 50000, "Comment", "Да",
        "Lenovo", "DNS",
        "Иванов И.И.", "Петров П.П.",
        "Windows 11", ""
    ]
    for i, val in enumerate(example, 1):
        ws.cell(row=2, column=i, value=val)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def _is_nan(value: Any) -> bool | ndarray[Any, dtype[bool[bool]]] | NDFrame:
    """Проверяет, является ли значение NaN (pandas/numpy)"""
    if value is None:
        return True
    if isinstance(value, float):
        return math.isnan(value) or math.isinf(value)
    if isinstance(value, (str, list, dict)) and len(value) == 0:
        return True
    return pd.isna(value)

def _clean_value(value: Any, field_name: str) -> Any:
    """Очищает значение от NaN и преобразует в нужный тип"""
    if _is_nan(value):
        return None

    # Обработка дат
    if field_name.endswith('_date'):
        if isinstance(value, pd.Timestamp):
            return value.date()
        elif isinstance(value, date):
            return value
        elif isinstance(value, str):
            try:
                return datetime.strptime(value, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                return None

    # Обработка булевых значений
    if field_name == 'affixed_inventory_id':
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ['да', 'yes', 'true', '1', 'y']
        return bool(value)

    # Обработка чисел
    if field_name == 'price':
        try:
            return int(float(value)) if value is not None else None
        except (ValueError, TypeError):
            return None

    # Для строк: возвращаем строку или None
    if isinstance(value, str):
        return value.strip() if value.strip() else None

    return str(value) if value is not None else None

async def parse_asset_import_excel(file_content: bytes) -> List[Dict[str, Any]]:
    """
    Парсит Excel файл импорта в список словарей.
    Ключи словаря будут соответствовать полям AssetExcelRow (snake_case).
    """
    df = pd.read_excel(io.BytesIO(file_content))
    rows = []

    # Создаем маппинг Заголовок -> ИмяПоля
    header_to_field = {header: field for field, header, _ in _get_excel_columns()}

    for _, row in df.iterrows():
        raw_dict = row.to_dict()
        mapped_dict = {}

        for excel_header, value in raw_dict.items():
            if excel_header in header_to_field:
                field_name = header_to_field[excel_header]
                mapped_dict[field_name] = _clean_value(value, field_name)

        # Валидация обязательных полей
        if (not mapped_dict.get('inventory_id') or
                not mapped_dict.get('serial_number') or
                not mapped_dict.get('name') or
                not mapped_dict.get('asset_type_name')):
            continue

        rows.append(mapped_dict)

    return rows