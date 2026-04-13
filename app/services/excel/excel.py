import io
import json
from datetime import date, datetime
from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Заголовки колонок в том порядке, как они будут в Excel
TEMPLATE_HEADERS = [
    "id", "caption", "inv_id", "serial_id", "type_id", "price",
    "status", "description", "delivery_date", "seller", "staff",
    "department", "parent_id", "fact_location", "source",
    "prepared_by", "checked_by", "user_info"
]

def create_template() -> bytes:
    """Создаёт Excel-шаблон с заголовками и примером данных (как в /xlsx/example.xlsx)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "IT Assets"  # Название листа как в шаблоне

    # Точный порядок колонок как в БД и в вашем файле
    TEMPLATE_HEADERS = [
        "id",
        "caption",
        "description",
        "inv_id",
        "serial_id",
        "status",
        "user_info",
        "seller",
        "price",
        "staff",
        "department",
        "fact_location",
        "source",
        "prepared_by",
        "checked_by",
        "type_id",
        "parent_id",
        "delivery_date",
        "deleted_at"
    ]

    # Стиль заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Заголовки
    for col, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    # Пример строки данных (как в вашем файле 1.xlsx)
    example = [
        17,                    # id
        "название",            # caption
        "Описание",            # description
        "инв",                 # inv_id
        "серийн",              # serial_id
        "Выдан",               # status
        '{"k": "v"}',          # user_info (JSON)
        "Продавец",            # seller
        100,                   # price
        "ответств",            # staff
        "отдел",               # department
        "место",               # fact_location
        "источник",            # source
        "подготовил",          # prepared_by
        "проверил",            # checked_by
        30,                    # type_id
        "",                    # parent_id (пусто)
        "2026-04-06",          # delivery_date
        ""                     # deleted_at (пусто)
    ]

    for col, val in enumerate(example, start=1):
        ws.cell(row=2, column=col, value=val)

    # Автоширина колонок
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # Сохраняем в bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def parse_excel_assets(file_content: bytes) -> List[Dict[str, Any]]:
    """
    Парсит загруженный Excel-файл и возвращает список словарей для создания активов.
    Возвращает также список ошибок, если данные невалидны.
    """
    wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active

    assets = []
    errors = []

    # Читаем заголовки из первой строки
    headers = [str(cell.value).strip().lower() for cell in ws[1]]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = dict(zip(headers, row))

        # Пропускаем полностью пустые строки
        if not any(row_data.values()):
            continue

        asset = {}
        row_errors = []

        try:
            # Обязательные поля
            for field in ["caption", "inv_id", "serial_id", "type_id", "price"]:
                val = row_data.get(field)
                if val is None or str(val).strip() == "":
                    row_errors.append(f"Поле '{field}' обязательно")
                else:
                    asset[field] = val

            if row_errors:
                errors.append(f"Строка {row_idx}: {'; '.join(row_errors)}")
                continue

            # Конвертация типов
            asset["type_id"] = int(float(asset["type_id"]))
            asset["price"] = float(asset["price"])

            # Опциональные поля
            asset["status"] = row_data.get("status") or "Приемка"
            # if asset["status"] not in VALID_STATUSES:
            #     errors.append(f"Строка {row_idx}: Неверный статус '{asset['status']}'")
            #     continue

            # Дата поставки
            delivery = row_data.get("delivery_date")
            if delivery:
                if isinstance(delivery, datetime):
                    asset["delivery_date"] = delivery.date().isoformat()
                elif isinstance(delivery, date):
                    asset["delivery_date"] = delivery.isoformat()
                else:
                    asset["delivery_date"] = str(delivery)[:10]  # Обрезаем до YYYY-MM-DD

            # Дата удаления
            deleted_at = row_data.get("deleted_at")
            if deleted_at:
                if isinstance(deleted_at, datetime):
                    asset["deleted_at"] = deleted_at.date().isoformat()
                elif isinstance(delivery, date):
                    asset["deleted_at"] = deleted_at.isoformat()
                else:
                    asset["deleted_at"] = str(deleted_at)[:10]  # Обрезаем до YYYY-MM-DD

            # Числовые опциональные
            parent = row_data.get("parent_id")
            asset["parent_id"] = int(float(parent)) if parent and str(parent).strip() else None

            # JSON поле user_info
            user_info = row_data.get("user_info")
            if user_info:
                import json
                try:
                    asset["user_info"] = json.loads(str(user_info)) if isinstance(user_info, str) else user_info
                except Exception:
                    asset["user_info"] = {"raw": str(user_info)}  # Сохраняем как есть, если не парсится
            else:
                asset["user_info"] = {}

            # Остальные строковые поля
            for field in ["description", "seller", "staff", "department",
                          "fact_location", "source", "prepared_by", "checked_by"]:
                val = row_data.get(field)
                asset[field] = str(val).strip() if val and str(val).strip() else None

            assets.append(asset)

        except (ValueError, TypeError) as e:
            errors.append(f"Строка {row_idx}: Ошибка конвертации данных — {str(e)}")
        except Exception as e:
            errors.append(f"Строка {row_idx}: Неизвестная ошибка — {str(e)}")

    wb.close()
    return assets, errors


def export_assets_to_excel(assets: List[Dict[str, Any]]) -> bytes:
    """
    Экспортирует список активов в Excel файл с точными именами полей из БД.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "IT Assets"

    # Точные имена полей из БД (как в screenshot)
    all_fields = [
        "id",
        "caption",
        "description",
        "inv_id",
        "serial_id",
        "status",
        "user_info",
        "seller",
        "price",
        "staff",
        "department",
        "fact_location",
        "source",
        "prepared_by",
        "checked_by",
        "type_id",
        "parent_id",
        "delivery_date",
        "deleted_at"
    ]

    # Стили
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Заголовки (используем точные имена из БД)
    for col, field in enumerate(all_fields, start=1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center")

    # Данные
    for row_idx, asset in enumerate(assets, start=2):
        for col, field in enumerate(all_fields, start=1):
            value = asset.get(field)

            # Форматирование значений
            if field == "user_info" and value:
                value = json.dumps(value, ensure_ascii=False) if isinstance(value, dict) else str(value)
            elif field == "deleted_at" and value:
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(value, date):
                    value = value.strftime("%Y-%m-%d")
            elif field == "delivery_date" and value:
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d")
                elif isinstance(value, date):
                    value = value.strftime("%Y-%m-%d")
            elif value is None:
                value = ""

            cell = ws.cell(row=row_idx, column=col, value=value)

            # Выделяем удалённые записи
            if field == "deleted_at" and value:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006")

    # Автоширина колонок
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Фильтры на заголовки
    ws.auto_filter.ref = ws.dimensions

    # Сохранение
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()