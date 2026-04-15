import io
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from app.schemas.catalog.CatalogExportSchemas import CatalogExportRow


def create_catalog_excel_file(catalog_data: List[CatalogExportRow]) -> bytes:
    """
    Создает Excel файл с данными каталога активов.
    Данные разделены на категории с цветовой кодировкой.

    Args:
        catalog_data: Список строк каталога для экспорта

    Returns:
        bytes: Содержимое Excel файла в байтах
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог активов"

    # Определяем стили
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill_category1 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Синий
    header_fill_category2 = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Зеленый
    header_fill_category3 = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")  # Оранжевый
    header_fill_category4 = PatternFill(start_color="7B7B7B", end_color="7B7B7B", fill_type="solid")  # Серый
    header_fill_category5 = PatternFill(start_color="9F4826", end_color="9F4826", fill_type="solid")  # Коричневый

    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Определяем колонки по категориям
    columns = [
        # Категория 1: Основная информация о записи
        ("catalog_id", "ID записи", header_fill_category1),

        # Категория 2: Класс оборудования
        ("class_name", "Класс оборудования", header_fill_category2),
        ("class_description", "Описание класса", header_fill_category2),

        # Категория 3: Модель оборудования
        ("model_name", "Модель оборудования", header_fill_category2),
        ("model_description", "Описание модели", header_fill_category2),
        ("model_is_active", "Модель активна", header_fill_category2),
        ("model_is_serial_required", "Требуется серийный номер", header_fill_category2),

        # Категория 4: Актив
        ("asset_inventory_id", "Инвентарный номер", header_fill_category3),
        ("asset_serial_number", "Серийный номер", header_fill_category3),
        ("asset_name", "Наименование актива", header_fill_category3),
        ("asset_status", "Статус актива", header_fill_category3),
        ("asset_type_domain", "Тип домена", header_fill_category3),
        ("asset_affixed_inventory_id", "Инв. номер наклеен", header_fill_category3),
        ("asset_info_storage_location", "Место хранения информации", header_fill_category3),
        ("asset_passwork", "Пароль/ключ", header_fill_category3),
        ("asset_date_issue", "Дата выдачи", header_fill_category3),
        ("asset_date_purchasing", "Дата покупки", header_fill_category3),
        ("asset_comment", "Комментарий к активу", header_fill_category3),
        ("asset_source", "Источник поступления", header_fill_category3),
        ("asset_seller", "Продавец/поставщик", header_fill_category3),
        ("asset_price", "Стоимость приобретения", header_fill_category3),

        # Категория 5: Владелец
        ("owner_name", "ФИО владельца", header_fill_category4),
        ("owner_email", "Email владельца", header_fill_category4),
        ("owner_department", "Отдел владельца", header_fill_category4),

        # Категория 6: Склад
        ("warehouse_name", "Название склада", header_fill_category5),
        ("warehouse_location_city", "Город склада", header_fill_category5),
        ("warehouse_location_address", "Адрес склада", header_fill_category5),

        # Категория 7: Гарантия
        ("warranty_end_date", "Дата окончания гарантии", header_fill_category1),

        # Категория 8: Аудит
        ("created_at", "Дата создания записи", header_fill_category1),
        ("created_by_name", "Создал (ФИО)", header_fill_category1),
        ("created_by_email", "Создал (Email)", header_fill_category1),
    ]

    # Создаем заголовки
    for col_idx, (field, header, fill) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = fill
        cell.border = thin_border

    # Заполняем данными
    for row_idx, row_data in enumerate(catalog_data, start=2):
        for col_idx, (field, _, _) in enumerate(columns, start=1):
            value = getattr(row_data, field)

            # Форматируем значения
            if isinstance(value, date):
                value = value.strftime("%Y-%m-%d")
            elif isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(value, bool):
                value = "Да" if value else "Нет"
            elif value is None:
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Автоширина для числовых колонок
            if field in ["catalog_id", "asset_price"]:
                cell.alignment = Alignment(horizontal='right', vertical='center')

    # Устанавливаем ширину колонок
    for col_idx, (_, header, _) in enumerate(columns, start=1):
        max_length = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        adjusted_width = min(max_length + 2, 50)  # Максимальная ширина 50
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Закрепляем первую строку (заголовки)
    ws.freeze_panes = "A2"

    # Сохраняем в буфер
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()